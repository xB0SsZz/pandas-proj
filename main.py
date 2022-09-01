import openpyxl as xl


def remove(xlsx_file):
    path = './' + xlsx_file.strip()
    wb = xl.load_workbook(path)
    sheet = wb.worksheets[0]
    sheet.delete_rows(1)
    sheet.delete_cols(23)
    wb.save('./' + xlsx_file.partition(".")[0] + "_croped.xlsx")
    read_config(wb)
    wb.save('./' + xlsx_file.partition(".")[0] + "_croped.xlsx")
    

def delete_rows(arg1, arg2, arg3, sheet, ws, column, j, counter):
    c = str(sheet.cell(j, column).value).replace(" ", "")
    if arg1 == c or arg1.strip() == c:
        counter += 1
        for k in range(1, sheet.max_column + 1):
            ws.cell(counter, k).value = sheet.cell(j, k).value
        sheet.delete_rows(j)
        delete_rows(arg1, arg2, arg3, sheet, ws, column, j, counter)

    if arg2 == c or arg2.strip() == c:
        counter += 1
        for k in range(1, sheet.max_column + 1):
            ws.cell(counter, k).value = sheet.cell(j, k).value
        sheet.delete_rows(j)
        delete_rows(arg1, arg2, arg3, sheet, ws, column, j, counter)

    if arg3 == c or arg3.strip() == c:
        counter += 1
        for k in range(1, sheet.max_column + 1):
            ws.cell(counter, k).value = sheet.cell(j, k).value
        sheet.delete_rows(j)
        delete_rows(arg1, arg2, arg3, sheet, ws, column, j, counter)
    return counter


def read_config(wb):
    sheet = wb.worksheets[0]
    with open("data_config_file.txt", "r") as file:
        for line in file:
            output_file = line.partition("|")[0]
            col = (line.partition("|")[2].partition("|")[0])
            arg1 = (line.partition("|")[2].partition("|")[2].partition("|")[0])
            arg2 = (line.partition("|")[2].partition("|")[2].partition("|")[2].partition("|")[0])
            arg3 = (line.partition("|")[2].partition("|")[2].partition("|")[2].partition("|")[2].partition("|")[0])
            col = col.replace(" ", "")
            arg1 = arg1.replace(" ", "")
            arg2 = arg2.replace(" ", "")
            arg3 = arg3.replace(" ", "")

            print(col + arg1 + arg2 + arg3 + "+")

            workb = xl.Workbook()
            ws = workb.worksheets[0]
            column = 0
            for i in range(1, sheet.max_column + 1):
                #print("Value: +" + sheet.cell(1, i).value.strip() + "+")
                c = str(sheet.cell(1, i).value).replace(" ", "")
                if c == col or c == col.strip():
                    column = i
                    
            if column == 0:
                print("No column with name " + col)
                return

            for i in range(1, sheet.max_column + 1):
                ws.cell(1, i).value = sheet.cell(1, i).value

            counter = 1
            for j in range(2, sheet.max_row + 1):
                counter = delete_rows(arg1, arg2, arg3, sheet, ws, column, j, counter)
                print(str(counter))

            workb.save("./" + output_file)


            



def main():
    remove('example.xlsx')

if __name__ == '__main__':
    main()